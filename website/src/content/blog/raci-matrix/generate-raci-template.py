#!/usr/bin/env python3
"""Generate the Excel templates offered on the RACI matrix post.

Lives beside the page it serves: `en.mdx` and `fr.mdx` in this folder,
published on /en/blog/raci-matrix and /fr/blog/raci-matrix.

Files produced, under `public/downloads/` and served on `/downloads/<file>`
    raci-matrix-template.xlsx   EN, linked from en.mdx
    matrice-raci-modele.xlsx    FR, linked from fr.mdx

Each workbook carries three tabs: an empty grid with a dropdown limited to the
four letters, a colour per letter and a counter flagging any filled row without
exactly one A, then the worked product launch example from the article, then the
four definitions.

These binaries are build artefacts committed to the repo. Edit the definitions
below and re-run this script rather than editing the files themselves.

    pip install openpyxl
    python3 src/content/blog/raci-matrix/generate-raci-template.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parents[4] / "public" / "downloads"

PURPLE = "9870F0"
GREY = "6B6B6B"
RED = "C0392B"

LETTERS = ["R", "A", "C", "I"]
LETTER_FILLS = {
    "R": "DCEBFB",  # doing the work
    "A": "E7DBFD",  # answering for it, Rolebase purple lightened
    "C": "FDECD2",  # consulted
    "I": "EDEDED",  # informed
}
BLANK_ROWS = 16


def _border():
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


def grid_sheet(ws, t, rows=()):
    """One RACI grid: title, header, then blank rows or a worked example."""
    ws.sheet_view.showGridLines = False
    people = t["people"]
    last_col = 2 + len(people)  # deliverable + people + the A counter

    ws.cell(1, 1, t["title"]).font = Font(bold=True, size=16, color=PURPLE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    for col, label in enumerate([t["deliverable"]] + people + [t["a_count"]], start=1):
        cell = ws.cell(3, col, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.cell(3, 1).alignment = Alignment(horizontal="left", vertical="center")

    first = 4
    last = 3 + max(len(rows), BLANK_ROWS if not rows else len(rows))
    for r in range(first, last + 1):
        row = rows[r - first] if r - first < len(rows) else None
        ws.cell(r, 1, row[0] if row else None).border = _border()
        for i in range(len(people)):
            cell = ws.cell(r, 2 + i, (row[1][i] or None) if row else None)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border()
        # Blank until the row names a deliverable, so an untouched grid stays quiet.
        counter = ws.cell(r, last_col,
                          f'=IF($A{r}="","",'
                          f'COUNTIF(B{r}:{get_column_letter(1 + len(people))}{r},"A"))')
        counter.alignment = Alignment(horizontal="center")
        counter.font = Font(color=GREY)
        counter.border = _border()

    cells = f"B{first}:{get_column_letter(1 + len(people))}{last}"
    dv = DataValidation(type="list", formula1='"R,A,C,I"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cells)
    for letter, colour in LETTER_FILLS.items():
        ws.conditional_formatting.add(cells, CellIsRule(
            operator="equal", formula=[f'"{letter}"'],
            fill=PatternFill(start_color=colour, end_color=colour, fill_type="solid")))

    counter_col = get_column_letter(last_col)
    counters = f"{counter_col}{first}:{counter_col}{last}"
    ws.conditional_formatting.add(counters, FormulaRule(
        formula=[f'AND($A{first}<>"",{counter_col}{first}<>1)'],
        font=Font(bold=True, color=RED)))

    ws.column_dimensions["A"].width = 42
    for i in range(len(people)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16
    ws.column_dimensions[get_column_letter(last_col)].width = 10

    ws.cell(last + 2, 1, t["note"]).font = Font(italic=True, size=9, color=GREY)
    ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=last_col)


def legend_sheet(ws, t):
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, t["legend_title"]).font = Font(bold=True, size=16, color=PURPLE)
    for col, label in enumerate((t["letter"], t["meaning"]), start=1):
        cell = ws.cell(3, col, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.border = _border()
    for i, letter in enumerate(LETTERS):
        row = 4 + i
        cell = ws.cell(row, 1, letter)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor=LETTER_FILLS[letter])
        cell.font = Font(bold=True)
        cell.border = _border()
        meaning = ws.cell(row, 2, t["legend"][letter])
        meaning.alignment = Alignment(wrap_text=True, vertical="center")
        meaning.border = _border()
        ws.row_dimensions[row].height = 32
    for i, rule in enumerate(t["rules"]):
        ws.cell(4 + len(LETTERS) + 1 + i, 1, rule).font = Font(color=GREY)
        ws.merge_cells(start_row=4 + len(LETTERS) + 1 + i, start_column=1,
                       end_row=4 + len(LETTERS) + 1 + i, end_column=2)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 78


def workbook(t):
    wb = Workbook()
    grid_sheet(wb.active, t)
    wb.active.title = t["tab_matrix"]
    grid_sheet(wb.create_sheet(t["tab_example"]), t["example"], t["example"]["rows"])
    legend_sheet(wb.create_sheet(t["tab_legend"]), t)
    return wb


FR_NOTE = "Modèle proposé par Rolebase, rolebase.io"
EN_NOTE = "Template by Rolebase, rolebase.io"

TEMPLATES = {
    "matrice-raci-modele.xlsx": {
        "title": "Matrice RACI",
        "deliverable": "Livrable",
        "people": ["Personne 1", "Personne 2", "Personne 3",
                   "Personne 4", "Personne 5", "Personne 6"],
        "a_count": "Nb de A",
        "note": FR_NOTE,
        "tab_matrix": "Matrice",
        "tab_example": "Exemple",
        "tab_legend": "Légende",
        "legend_title": "Les quatre lettres",
        "letter": "Lettre",
        "meaning": "Signification",
        "legend": {
            "R": "Responsible, la ou les personnes qui réalisent le travail. "
                 "Plusieurs R sur une ligne, c'est normal.",
            "A": "Accountable, la seule personne qui répond du résultat. "
                 "Un seul A par ligne, toujours.",
            "C": "Consulted, ceux dont l'avis est nécessaire avant la livraison, "
                 "dans un échange à double sens.",
            "I": "Informed, ceux que l'on prévient une fois le travail terminé, "
                 "dans un seul sens.",
        },
        "rules": [
            "La colonne « Nb de A » reste vide tant que la ligne ne nomme pas de livrable, "
            "puis passe en rouge si cette ligne compte zéro ou plusieurs A.",
            "Une case vide est une vraie réponse : cette personne reste en dehors de ce livrable.",
            "Gardez les lignes au niveau du livrable plutôt qu'à celui de la tâche. "
            "Douze à vingt lignes suffisent.",
            "Renommez les colonnes « Personne 1 » à « Personne 6 » avec les personnes "
            "ou les rôles de votre projet.",
        ],
        "example": {
            "title": "Exemple rempli, lancement d'un produit",
            "deliverable": "Livrable",
            "people": ["Chef de produit", "Lead technique", "Marketing",
                       "Resp. commercial", "Direction"],
            "a_count": "Nb de A",
            "note": FR_NOTE,
            "rows": [
                ("Fixer la date de lancement", ["A", "C", "C", "C", "I"]),
                ("Geler le périmètre fonctionnel", ["A", "R", "I", "", "I"]),
                ("Rédiger les notes de version", ["C", "R", "A", "I", ""]),
                ("Construire la page de lancement", ["C", "", "A", "I", ""]),
                ("Fixer le prix de lancement", ["C", "", "C", "C", "A"]),
                ("Briefer l'équipe commerciale", ["I", "", "R", "A", ""]),
                ("Annoncer aux clients", ["I", "", "R", "I", "A"]),
            ],
        },
    },
    "raci-matrix-template.xlsx": {
        "title": "RACI matrix",
        "deliverable": "Deliverable",
        "people": ["Person 1", "Person 2", "Person 3",
                   "Person 4", "Person 5", "Person 6"],
        "a_count": "A count",
        "note": EN_NOTE,
        "tab_matrix": "Matrix",
        "tab_example": "Example",
        "tab_legend": "Legend",
        "legend_title": "The four letters",
        "letter": "Letter",
        "meaning": "Meaning",
        "legend": {
            "R": "Responsible, the people who do the work. "
                 "Several R's on one row is normal.",
            "A": "Accountable, the one person who answers for the result. "
                 "One A per row, always.",
            "C": "Consulted, the people whose input is needed before delivery, "
                 "in a two-way exchange.",
            "I": "Informed, the people told once the work is done, "
                 "in one direction only.",
        },
        "rules": [
            "The A count column stays blank until the row names a deliverable, then turns red "
            "if that row carries zero or several A's.",
            "An empty cell is a real answer. That person stays out of this deliverable.",
            "Keep the rows at the level of a deliverable rather than a to-do item. "
            "Twelve to twenty rows are enough.",
            "Rename the Person 1 to Person 6 columns with the people or roles "
            "on your project.",
        ],
        "example": {
            "title": "Worked example, launching a product",
            "deliverable": "Deliverable",
            "people": ["Product manager", "Engineering lead", "Marketing",
                       "Sales lead", "CEO"],
            "a_count": "A count",
            "note": EN_NOTE,
            "rows": [
                ("Set the launch date", ["A", "C", "C", "C", "I"]),
                ("Freeze the feature scope", ["A", "R", "I", "", "I"]),
                ("Write the release notes", ["C", "R", "A", "I", ""]),
                ("Build the landing page", ["C", "", "A", "I", ""]),
                ("Set the launch price", ["C", "", "C", "C", "A"]),
                ("Brief the sales team", ["I", "", "R", "A", ""]),
                ("Announce to customers", ["I", "", "R", "I", "A"]),
            ],
        },
    },
}

if __name__ == "__main__":
    OUT.mkdir(parents=True, exist_ok=True)
    for filename, spec in TEMPLATES.items():
        workbook(spec).save(OUT / filename)
        print(f"  {filename}")
